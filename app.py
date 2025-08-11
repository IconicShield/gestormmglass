# app.py

# A ORDEM AQUI É CRUCIAL PARA O EVENTLET
# import eventlet
# eventlet.monkey_patch()

# AGORA, importe todo o resto
import os
from dotenv import load_dotenv
from flask import Flask, Response, render_template, request, redirect, url_for, send_from_directory, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from datetime import datetime
from sqlalchemy.exc import IntegrityError
from sqlalchemy.pool import NullPool
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_bcrypt import Bcrypt
from sqlalchemy import or_, cast
from functools import wraps
import click
import io
import openpyxl
from getpass import getpass
# from flask_socketio import SocketIO
import fitz
import re
import io
import openpyxl
import time

# --- INÍCIO DA CORREÇÃO ESTRUTURAL ---

# 1. Carregue as variáveis de ambiente primeiro
load_dotenv()

# 2. Crie as instâncias das extensões SEM o app
db = SQLAlchemy()
bcrypt = Bcrypt()
login_manager = LoginManager()
# socketio = SocketIO()

# 3. Crie a instância do app Flask
app = Flask(__name__)
app.jinja_env.add_extension('jinja2.ext.do')

# app.py

# --- BLOCO DE CONFIGURAÇÃO DO BANCO DE DADOS (VERSÃO CORRIGIDA) ---

# Define o diretório base
basedir = os.path.abspath(os.path.dirname(__file__))

# Carrega a variável de ambiente DATABASE_URL do arquivo .env
database_url = os.environ.get('DATABASE_URL')

# Configura o app com base na variável
app.config['SECRET_KEY'] = 'uma-chave-secreta-muito-dificil-de-adivinhar'

# Verifica se a variável do Supabase existe para construir a URI
if database_url:
    # Substitui o prefixo para compatibilidade com SQLAlchemy
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)
    
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
    print("--- Conectando ao banco de dados Supabase ---")
else:
    # Caso contrário, usa o banco de dados local SQLite
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'database.db')
    print("--- Usando banco de dados local SQLite ---")

app.config['UPLOAD_FOLDER'] = os.path.join(basedir, 'uploads')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'poolclass': NullPool}
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# --- FIM DO BLOCO CORRIGIDO ---

# 5. INICIALIZE as extensões COM o app
db.init_app(app)
bcrypt.init_app(app)
login_manager.init_app(app)
# socketio.init_app(app, async_mode='eventlet')

# 6. Configure o LoginManager
login_manager.login_view = 'login'
login_manager.login_message = "Por favor, faça login para aceder a esta página."
login_manager.login_message_category = "info"

# --- FIM DA CORREÇÃO ESTRUTURAL ---


class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password_hash = db.Column(db.String(150), nullable=False)
    is_admin = db.Column(db.Boolean, default=False, nullable=False)

class Cliente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    numero_cliente = db.Column(db.Integer, unique=True, nullable=False)
    nome = db.Column(db.String(150), nullable=False)
    telefone = db.Column(db.String(20), nullable=True)
    
    # --- NOVOS CAMPOS ---
    tipo_pessoa = db.Column(db.String(10), nullable=False) # Física ou Jurídica
    cpf_cnpj = db.Column(db.String(20), nullable=True, unique=True)
    como_conheceu = db.Column(db.String(100), nullable=True)
    
    # Endereço (cadastro completo)
    rua = db.Column(db.String(200), nullable=True)
    numero_endereco = db.Column(db.String(20), nullable=True)
    complemento = db.Column(db.String(100), nullable=True)
    bairro = db.Column(db.String(100), nullable=True)
    cidade = db.Column(db.String(100), nullable=True)
    uf = db.Column(db.String(2), nullable=True)
    cep = db.Column(db.String(10), nullable=True)
    
    observacoes = db.Column(db.Text, nullable=True)

    def __repr__(self):
        return self.nome

class Anexo(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(200), nullable=False)
    entrada_id = db.Column(db.Integer, db.ForeignKey('entrada.id'), nullable=False)

class Entrada(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tipo = db.Column(db.String(50), nullable=False)
    numero_pedido = db.Column(db.Integer, unique=True, nullable=False)
    data_registro = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    
    # --- INÍCIO DA ALTERAÇÃO CRÍTICA ---
    # Ligação formal com a tabela de Clientes
    cliente_id = db.Column(db.Integer, db.ForeignKey('cliente.id'), nullable=True)
    cliente = db.relationship('Cliente', backref='entradas')
    
    # Campo para guardar o nome se o cliente não for cadastrado
    cliente_nome_temp = db.Column(db.String(150), nullable=True)
    # --- FIM DA ALTERAÇÃO CRÍTICA ---
    obra = db.Column(db.String(200), nullable=True)
    status = db.Column(db.String(50), nullable=False, default='Não iniciado')
    descricao = db.Column(db.Text, nullable=False)
    observacoes = db.Column(db.Text, nullable=True)
    arquivado = db.Column(db.Boolean, default=False, nullable=False)
    anexos = db.relationship('Anexo', backref='entrada', lazy=True, cascade="all, delete-orphan")

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin:
            flash('Acesso negado. Apenas administradores podem aceder a esta página.', 'danger')
            return redirect(url_for('inicio'))
        return f(*args, **kwargs)
    return decorated_function


def get_dashboard_data():
    pedidos_query = Entrada.query.filter_by(tipo='Pedido', arquivado=False)
    pedidos_dashboard = {
        'total': pedidos_query.count(),
        'nao_iniciado': pedidos_query.filter_by(status='Não iniciado').count(),
        'em_andamento': pedidos_query.filter_by(status='Em andamento').count(),
        'concluido': pedidos_query.filter_by(status='Concluído').count()
    }
    orcamentos_query = Entrada.query.filter_by(tipo='Orçamento', arquivado=False)
    orcamentos_dashboard = {
        'total': orcamentos_query.count(),
        'nao_iniciado': orcamentos_query.filter_by(status='Não iniciado').count(),
        'em_andamento': orcamentos_query.filter_by(status='Em andamento').count(),
        'concluido': orcamentos_query.filter_by(status='Concluído').count()
    }
    return {'pedidos': pedidos_dashboard, 'orcamentos': orcamentos_dashboard}

@app.template_filter('format_phone')
def format_phone_filter(s):
    """Formata um número de telefone para (xx) x xxxx-xxxx."""
    if not s or not s.isdigit() or len(s) != 11:
        return s # Retorna o original se não for um número de 11 dígitos
    return f"({s[0:2]}) {s[2]} {s[3:7]}-{s[7:11]}"

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('inicio'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        if user and bcrypt.check_password_hash(user.password_hash, password):
            login_user(user)
            return redirect(url_for('inicio'))
        else:
            flash('Login sem sucesso. Verifique o utilizador e a senha.', 'danger')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logout efetuado com sucesso.', 'success')
    return redirect(url_for('login'))

@app.route('/')
@login_required
def inicio():
    """Renderiza a nova página de início."""
    return render_template('inicio.html')

@app.route('/painel')
@login_required
def painel_controle():
    search_query = request.args.get('q', '')
    selected_status = request.args.get('status', '')

    # Usamos outerjoin para incluir entradas mesmo que não tenham cliente_id
    pedidos_base_query = Entrada.query.outerjoin(Cliente).filter(
        Entrada.tipo == 'Pedido', 
        Entrada.arquivado == False
    )
    orcamentos_base_query = Entrada.query.outerjoin(Cliente).filter(
        Entrada.tipo == 'Orçamento', 
        Entrada.arquivado == False
    )

    if search_query:
        search_term = f"%{search_query}%"
        # O filtro de busca agora procura em 4 lugares diferentes
        search_filter = or_(
            cast(Entrada.numero_pedido, db.String).ilike(search_term),
            Entrada.descricao.ilike(search_term),
            Cliente.nome.ilike(search_term),
            Entrada.cliente_nome_temp.ilike(search_term)
        )
        pedidos_base_query = pedidos_base_query.filter(search_filter)
        orcamentos_base_query = orcamentos_base_query.filter(search_filter)
    
    if selected_status:
        pedidos_base_query = pedidos_base_query.filter(Entrada.status == selected_status)
        orcamentos_base_query = orcamentos_base_query.filter(Entrada.status == selected_status)
        
    pedidos = pedidos_base_query.order_by(Entrada.numero_pedido).all()
    orcamentos = orcamentos_base_query.order_by(Entrada.numero_pedido).all()
    
    dashboard_data = get_dashboard_data()
    
    return render_template('painel_controle.html', dashboard=dashboard_data, pedidos=pedidos, orcamentos=orcamentos, search_query=search_query, selected_status=selected_status)

@app.route('/novo', methods=['GET', 'POST'])
@login_required
def nova_entrada():
    if request.method == 'POST':
        tipo_entrada = request.form.get('tipo')
        numero_pedido_str = request.form.get('numero_pedido')

        if tipo_entrada == 'Pedido' and not numero_pedido_str:
            flash('O N° da Entrada é obrigatório para Pedidos.', 'danger')
            return render_template('nova_entrada.html', form_data=request.form)

        if tipo_entrada == 'Orçamento' and not numero_pedido_str:
            numero_pedido = int(datetime.now().timestamp())
        else:
            if not numero_pedido_str.isdigit():
                flash('O N° de entrada deve conter apenas números.', 'danger')
                return render_template('nova_entrada.html', form_data=request.form)
            numero_pedido = int(numero_pedido_str)

        if Entrada.query.filter_by(numero_pedido=numero_pedido).first():
            flash(f'O N° de entrada {numero_pedido} já existe. Tente outro.', 'danger')
            return render_template('nova_entrada.html', form_data=request.form)

        # --- LÓGICA DO NOVO CLIENTE ---
        cliente_id = request.form.get('cliente_id')
        cliente_nome = request.form.get('cliente_nome')

        nova_entrada_obj = Entrada(
            tipo=tipo_entrada,
            numero_pedido=numero_pedido,
            status=request.form.get('status'),
            obra=request.form.get('obra'),
            descricao=request.form.get('descricao'),
            observacoes=request.form.get('observacoes')
        )

        if cliente_id: # Se um cliente existente foi selecionado
            nova_entrada_obj.cliente_id = int(cliente_id)
        else: # Se um novo nome foi digitado
            nova_entrada_obj.cliente_nome_temp = cliente_nome
        # --- FIM DA LÓGICA ---

        db.session.add(nova_entrada_obj)
        uploaded_files = request.files.getlist('anexos')
        for ficheiro in uploaded_files:
            if ficheiro and ficheiro.filename != '':
                anexo_filename = secure_filename(ficheiro.filename)
                ficheiro.save(os.path.join(app.config['UPLOAD_FOLDER'], anexo_filename))
                novo_anexo = Anexo(filename=anexo_filename, entrada=nova_entrada_obj)
                db.session.add(novo_anexo)
        db.session.commit()
        global last_update_time
        last_update_time = datetime.now()
        flash(f"{nova_entrada_obj.tipo} criado com sucesso!", 'success')
        return redirect(url_for('painel_controle'))

    return render_template('nova_entrada.html', form_data={})

@app.route('/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_entrada(id):
    entrada = Entrada.query.get_or_404(id)
    if request.method == 'POST':
        # Atualiza os campos simples
        entrada.tipo = request.form.get('tipo')
        entrada.numero_pedido = int(request.form.get('numero_pedido'))
        entrada.status = request.form.get('status')
        entrada.obra = request.form.get('obra')
        entrada.descricao = request.form.get('descricao')
        entrada.observacoes = request.form.get('observacoes')

        # --- NOVA LÓGICA PARA ATUALIZAR O CLIENTE ---
        cliente_id = request.form.get('cliente_id')
        cliente_nome = request.form.get('cliente_nome')

        if cliente_id:
            # Se um cliente existente foi selecionado, vincula o ID
            entrada.cliente_id = int(cliente_id)
            entrada.cliente_nome_temp = None # Limpa o nome temporário
        else:
            # Se um novo nome foi digitado, guarda como temporário
            entrada.cliente_id = None
            entrada.cliente_nome_temp = cliente_nome
        # --- FIM DA NOVA LÓGICA ---

        # Lógica para adicionar novos anexos (se houver)
        uploaded_files = request.files.getlist('anexos')
        for ficheiro in uploaded_files:
            if ficheiro and ficheiro.filename != '':
                anexo_filename = secure_filename(ficheiro.filename)
                ficheiro.save(os.path.join(app.config['UPLOAD_FOLDER'], anexo_filename))
                novo_anexo = Anexo(filename=anexo_filename, entrada=entrada)
                db.session.add(novo_anexo)
                
        db.session.commit()
        global last_update_time
        last_update_time = datetime.now()
        flash(f'{entrada.tipo} atualizado com sucesso!', 'success')
        return redirect(url_for('painel_controle'))
        
    return render_template('editar_entrada.html', entrada=entrada)

@app.route('/excluir-anexo/<int:anexo_id>', methods=['GET', 'POST'])
@login_required
def excluir_anexo(anexo_id):
    anexo = Anexo.query.get_or_404(anexo_id)
    entrada_id = anexo.entrada_id
    try:
        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], anexo.filename))
    except FileNotFoundError:
        pass
    db.session.delete(anexo)
    db.session.commit()
    # socketio.emit('update_data')
    flash('Anexo excluído com sucesso.', 'success')
    return redirect(url_for('editar_entrada', id=entrada_id))

@app.route('/excluir/<int:id>', methods=['POST'])
@login_required
def excluir_entrada(id):
    global last_update_time
    entrada_a_excluir = Entrada.query.get_or_404(id)
    for anexo in entrada_a_excluir.anexos:
        try:
            os.remove(os.path.join(app.config['UPLOAD_FOLDER'], anexo.filename))
        except FileNotFoundError:
            pass
    db.session.delete(entrada_a_excluir)
    db.session.commit()
    # socketio.emit('update_data')
    last_update_time = datetime.now()
    tipo_entrada = entrada_a_excluir.tipo
    if entrada_a_excluir.arquivado:
        flash(f'{tipo_entrada} arquivado foi excluído permanentemente!', 'danger')
        return redirect(url_for('pedidos_arquivados'))
    flash(f'{tipo_entrada} foi excluído com sucesso!', 'danger')
    return redirect(url_for('painel_controle'))

# Variável global para controlar atualizações
last_update_time = datetime.now()

@app.route('/atualizar-status/<int:id>', methods=['POST'])
@login_required
def atualizar_status(id):
    global last_update_time
    entrada = Entrada.query.get_or_404(id)
    data = request.get_json()
    novo_status = data.get('status')
    if novo_status in ['Não iniciado', 'Em andamento', 'Concluído']:
        entrada.status = novo_status
        db.session.commit()
        last_update_time = datetime.now()
        novos_dados_dashboard = get_dashboard_data()
        return jsonify({'success': True, 'message': 'Status atualizado com sucesso!', 'dashboard': novos_dados_dashboard})
    return jsonify({'success': False, 'message': 'Status inválido.'}), 400

@app.route('/stream')
@login_required
def stream():
    """Server-Sent Events para atualizações em tempo real"""
    def event_stream():
        global last_update_time
        client_last_update = datetime.now()
        
        while True:
            # Verifica se houve atualizações
            if last_update_time > client_last_update:
                dashboard_data = get_dashboard_data()
                import json
                yield f"data: {json.dumps(dashboard_data)}\n\n"
                client_last_update = datetime.now()
            
            # Aguarda 3 segundos antes da próxima verificação
            import time
            time.sleep(3)
    
    return Response(event_stream(), mimetype="text/event-stream", headers={
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive'
    })

@app.route('/converter/<int:id>', methods=['POST'])
@login_required
def converter_para_pedido(id):
    orcamento = Entrada.query.get_or_404(id)
    if orcamento.tipo == 'Orçamento':
        orcamento.tipo = 'Pedido'
        orcamento.status = 'Não iniciado'
        db.session.commit()
        global last_update_time
        last_update_time = datetime.now()
        flash(f"Orçamento #{orcamento.numero_pedido} foi convertido em Pedido com sucesso!", 'success')
    else:
        flash('Esta entrada já é um Pedido.', 'warning')
    return redirect(url_for('painel_controle'))

@app.route('/arquivar/<int:id>', methods=['POST'])
@login_required
def arquivar_entrada(id):
    entrada = Entrada.query.get_or_404(id)
    entrada.arquivado = True
    db.session.commit()
    global last_update_time
    last_update_time = datetime.now()
    flash(f'{entrada.tipo} #{entrada.numero_pedido} foi arquivado com sucesso.', 'success')
    return redirect(url_for('painel_controle'))

@app.route('/bulk-action', methods=['POST'])
@login_required
def bulk_action():
    """Processa ações em lote (arquivar ou excluir múltiplas entradas)."""
    try:
        data = request.get_json()
        action = data.get('action')
        ids = data.get('ids', [])
        entry_type = data.get('type')
        
        if not action or not ids:
            return jsonify({'success': False, 'message': 'Parâmetros inválidos'}), 400
        
        # Converte IDs para inteiros
        try:
            ids = [int(id_str) for id_str in ids]
        except ValueError:
            return jsonify({'success': False, 'message': 'IDs inválidos'}), 400
        
        # Busca as entradas
        entradas = Entrada.query.filter(Entrada.id.in_(ids)).all()
        
        if len(entradas) != len(ids):
            return jsonify({'success': False, 'message': 'Algumas entradas não foram encontradas'}), 404
        
        if action == 'archive':
            # Arquiva as entradas selecionadas
            for entrada in entradas:
                entrada.arquivado = True
            db.session.commit()
            # socketio.emit('update_data')
            return jsonify({
                'success': True, 
                'message': f'{len(entradas)} entrada(s) arquivada(s) com sucesso'
            })
            
        elif action == 'delete':
            # Exclui as entradas selecionadas
            for entrada in entradas:
                # Remove anexos físicos
                for anexo in entrada.anexos:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], anexo.filename))
                    except FileNotFoundError:
                        pass
                # Remove a entrada do banco
                db.session.delete(entrada)
            
            db.session.commit()
            global last_update_time
            last_update_time = datetime.now()
            return jsonify({
                'success': True, 
                'message': f'{len(entradas)} entrada(s) excluída(s) com sucesso'
            })
        
        else:
            return jsonify({'success': False, 'message': 'Ação não reconhecida'}), 400
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/bulk-action-archived', methods=['POST'])
@login_required
def bulk_action_archived():
    """Processa ações em lote para entradas arquivadas (restaurar ou excluir permanentemente)."""
    try:
        data = request.get_json()
        action = data.get('action')
        ids = data.get('ids', [])
        
        if not action or not ids:
            return jsonify({'success': False, 'message': 'Parâmetros inválidos'}), 400
        
        # Converte IDs para inteiros
        try:
            ids = [int(id_str) for id_str in ids]
        except ValueError:
            return jsonify({'success': False, 'message': 'IDs inválidos'}), 400
        
        # Busca as entradas arquivadas
        entradas = Entrada.query.filter(Entrada.id.in_(ids), Entrada.arquivado == True).all()
        
        if len(entradas) != len(ids):
            return jsonify({'success': False, 'message': 'Algumas entradas não foram encontradas'}), 404
        
        if action == 'restore':
            # Restaura as entradas selecionadas
            for entrada in entradas:
                entrada.arquivado = False
            db.session.commit()
            return jsonify({
                'success': True, 
                'message': f'{len(entradas)} entrada(s) restaurada(s) com sucesso'
            })
            
        elif action == 'delete':
            # Exclui permanentemente as entradas selecionadas
            for entrada in entradas:
                # Remove anexos físicos
                for anexo in entrada.anexos:
                    try:
                        os.remove(os.path.join(app.config['UPLOAD_FOLDER'], anexo.filename))
                    except FileNotFoundError:
                        pass
                # Remove a entrada do banco
                db.session.delete(entrada)
            
            db.session.commit()
            return jsonify({
                'success': True, 
                'message': f'{len(entradas)} entrada(s) excluída(s) permanentemente com sucesso'
            })
        
        else:
            return jsonify({'success': False, 'message': 'Ação não reconhecida'}), 400
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/bulk-action-clientes', methods=['POST'])
@login_required
def bulk_action_clientes():
    """Processa ações em lote para clientes (excluir múltiplos clientes)."""
    try:
        data = request.get_json()
        action = data.get('action')
        ids = data.get('ids', [])
        
        if not action or not ids:
            return jsonify({'success': False, 'message': 'Parâmetros inválidos'}), 400
        
        # Converte IDs para inteiros
        try:
            ids = [int(id_str) for id_str in ids]
        except ValueError:
            return jsonify({'success': False, 'message': 'IDs inválidos'}), 400
        
        # Busca os clientes
        clientes = Cliente.query.filter(Cliente.id.in_(ids)).all()
        
        if len(clientes) != len(ids):
            return jsonify({'success': False, 'message': 'Alguns clientes não foram encontrados'}), 404
        
        if action == 'delete':
            # Verifica se algum cliente tem entradas associadas
            clientes_com_entradas = []
            for cliente in clientes:
                if cliente.entradas:
                    clientes_com_entradas.append(cliente.nome)
            
            if clientes_com_entradas:
                return jsonify({
                    'success': False, 
                    'message': f'Não é possível excluir os seguintes clientes pois possuem pedidos/orçamentos associados: {", ".join(clientes_com_entradas)}'
                }), 400
            
            # Exclui os clientes selecionados
            for cliente in clientes:
                db.session.delete(cliente)
            
            db.session.commit()
            return jsonify({
                'success': True, 
                'message': f'{len(clientes)} cliente(s) excluído(s) com sucesso'
            })
        
        else:
            return jsonify({'success': False, 'message': 'Ação não reconhecida'}), 400
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Erro interno: {str(e)}'}), 500

@app.route('/desarquivar/<int:id>', methods=['POST'])
@login_required
def desarquivar_entrada(id):
    entrada = Entrada.query.get_or_404(id)
    entrada.arquivado = False
    db.session.commit()
    global last_update_time
    last_update_time = datetime.now()
    flash(f'{entrada.tipo} #{entrada.numero_pedido} foi restaurado com sucesso.', 'success')
    return redirect(url_for('pedidos_arquivados'))

@app.route('/arquivados')
@login_required
def pedidos_arquivados():
    """Exibe a página de pedidos e orçamentos arquivados com paginação e filtros (30 itens por página)."""
    # Parâmetros de filtro
    search_cliente = request.args.get('search_cliente', '').strip()
    filter_status = request.args.get('filter_status', '')
    filter_tipo = request.args.get('filter_tipo', '')
    page_pedidos = request.args.get('pedidos_page', 1, type=int)
    page_orcamentos = request.args.get('orcamentos_page', 1, type=int)
    per_page = 30  # Limite de 30 itens por página

    # Query base para entradas arquivadas
    base_query = Entrada.query.filter_by(arquivado=True)
    
    # Aplicar filtro por tipo se especificado
    if filter_tipo:
        base_query = base_query.filter(Entrada.tipo == filter_tipo)
    
    # Aplicar filtro por cliente se especificado
    if search_cliente:
        base_query = base_query.join(Cliente, Entrada.cliente_id == Cliente.id, isouter=True).filter(
            db.or_(
                Cliente.nome.ilike(f'%{search_cliente}%'),
                Entrada.cliente_nome_temp.ilike(f'%{search_cliente}%')
            )
        )
    
    # Aplicar filtro por status se especificado
    if filter_status:
        base_query = base_query.filter(Entrada.status == filter_status)
    
    # Query para pedidos arquivados - criar uma nova query baseada na base_query
    if not filter_tipo or filter_tipo == 'Pedido':
        pedidos_query = Entrada.query.filter_by(arquivado=True).filter(Entrada.tipo == 'Pedido')
        # Aplicar os mesmos filtros da base_query
        if search_cliente:
            pedidos_query = pedidos_query.join(Cliente, Entrada.cliente_id == Cliente.id, isouter=True).filter(
                db.or_(
                    Cliente.nome.ilike(f'%{search_cliente}%'),
                    Entrada.cliente_nome_temp.ilike(f'%{search_cliente}%')
                )
            )
        if filter_status:
            pedidos_query = pedidos_query.filter(Entrada.status == filter_status)
    else:
        pedidos_query = Entrada.query.filter_by(id=-1)  # Query vazia se filtro não for Pedido
    
    # Query para orçamentos arquivados - criar uma nova query baseada na base_query
    if not filter_tipo or filter_tipo == 'Orçamento':
        orcamentos_query = Entrada.query.filter_by(arquivado=True).filter(Entrada.tipo == 'Orçamento')
        # Aplicar os mesmos filtros da base_query
        if search_cliente:
            orcamentos_query = orcamentos_query.join(Cliente, Entrada.cliente_id == Cliente.id, isouter=True).filter(
                db.or_(
                    Cliente.nome.ilike(f'%{search_cliente}%'),
                    Entrada.cliente_nome_temp.ilike(f'%{search_cliente}%')
                )
            )
        if filter_status:
            orcamentos_query = orcamentos_query.filter(Entrada.status == filter_status)
    else:
        orcamentos_query = Entrada.query.filter_by(id=-1)  # Query vazia se filtro não for Orçamento
    
    # Paginação para pedidos
    pedidos_paginados = pedidos_query.order_by(Entrada.data_registro.desc()).paginate(
        page=page_pedidos, per_page=per_page, error_out=False
    )
    
    # Paginação para orçamentos
    orcamentos_paginados = orcamentos_query.order_by(Entrada.data_registro.desc()).paginate(
        page=page_orcamentos, per_page=per_page, error_out=False
    )

    return render_template('pedidos_arquivados.html', 
                         pedidos=pedidos_paginados, 
                         orcamentos=orcamentos_paginados,
                         search_cliente=search_cliente,
                         filter_status=filter_status,
                         filter_tipo=filter_tipo,
                         pedidos_page=page_pedidos,
                         orcamentos_page=page_orcamentos)


@app.route('/gerir_usuarios', methods=['GET', 'POST'])
@login_required
@admin_required
def gerir_usuarios():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        is_admin = 'is_admin' in request.form
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash('Este nome de utilizador já existe.', 'danger')
        else:
            hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
            new_user = User(username=username, password_hash=hashed_password, is_admin=is_admin)
            db.session.add(new_user)
        db.session.commit()
        # socketio.emit('update_data')
        flash(f'Utilizador {username} criado com sucesso!', 'success')
        return redirect(url_for('gerir_usuarios'))
    users = User.query.all()
    return render_template('gerir_usuarios.html', users=users)


@app.route('/excluir_usuario/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def excluir_usuario(user_id):
    user_to_delete = User.query.get_or_404(user_id)
    if user_to_delete.id == current_user.id:
        flash('Não pode excluir a sua própria conta de administrador.', 'danger')
        return redirect(url_for('gerir_usuarios'))
    db.session.delete(user_to_delete)
    db.session.commit()
    # socketio.emit('update_data')
    flash(f'Utilizador {user_to_delete.username} excluído com sucesso.', 'success')
    return redirect(url_for('gerir_usuarios'))
    

@app.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    response = send_from_directory(app.config['UPLOAD_FOLDER'], filename)
    # Adiciona o cabeçalho que sugere a visualização "inline"
    response.headers['Content-Disposition'] = f'inline; filename="{filename}"'
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    return response

@app.route('/relatorio-romaneio', methods=['GET', 'POST'])
@login_required
def relatorio_romaneio():
    if request.method == 'POST':
        if 'pdf_file' not in request.files:
            flash('Nenhum ficheiro selecionado.', 'warning')
            return redirect(request.url)
        
        ficheiro = request.files['pdf_file']
        
        if ficheiro.filename == '':
            flash('Nenhum ficheiro selecionado.', 'warning')
            return redirect(request.url)
        
        if ficheiro and ficheiro.filename.lower().endswith('.pdf'):
            try:
                pdf_bytes = ficheiro.read()
                texto_completo = ""
                with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
                    for page in doc:
                        # Extrai o texto mantendo algum layout básico
                        texto_completo += page.get_text()

                # --- LÓGICA DE EXTRAÇÃO AVANÇADA ---
                
                relatorios_extraidos = []
                # Divide o documento em blocos de pedidos, usando "Pedido Pedido Cli." como separador
                blocos_pedidos = re.split(r'Pedido\s+Pedido Cli\.', texto_completo, flags=re.IGNORECASE)

                for bloco in blocos_pedidos[1:]: # Ignora o que vem antes do primeiro pedido
                    
                    # 1. Extrai o NOME DO CLIENTE do texto que vem ANTES do cabeçalho do pedido
                    # Padrão: Procura por uma linha com letras maiúsculas que termina com " - VID."
                    cliente_match = re.search(r'\n([A-Z\s\d\.\-]+-\s*VID\..*?)\n', bloco)
                    nome_cliente = cliente_match.group(1).strip() if cliente_match else "Cliente não encontrado"

                    # 2. Extrai os dados do cabeçalho do pedido
                    # Padrão: Procura pela linha de valores que vem logo após o cabeçalho
                    padrao_cabecalho = re.search(r'Tipo\s+Funcionário\s+Data Pedido\s+Data Entrega\s+Peso\s+m²\s+Total\n(.*?)\n', bloco, re.IGNORECASE | re.DOTALL)
                    dados_cabecalho = {}
                    if padrao_cabecalho:
                        # Divide a linha de valores por múltiplos espaços
                        valores = re.split(r'\s{2,}', padrao_cabecalho.group(1).strip())
                        if len(valores) >= 9:
                            dados_cabecalho = {
                                'pedido': valores[0], 'pedido_cli': valores[1], 'tipo': valores[2],
                                'funcionario': valores[3], 'data_pedido': valores[4], 'data_entrega': valores[5],
                                'peso': valores[6], 'm2': valores[7], 'total': valores[8],
                                'cliente': nome_cliente # Adiciona o cliente que encontrámos
                            }

                    # 3. Extrai a tabela de produtos
                    produtos = []
                    # Padrão para encontrar uma linha de produto, que pode ser complexa
                    padrao_produtos = re.compile(r'(\d{4,})\s+(.*?)\s+OS:(\d+)\s+(\d+x\d+)\s+(\d+)\s+([\d,]+)', re.DOTALL)
                    
                    seccao_produtos_match = re.search(r'Cod\s+Produto\s+LarguraxAltura(.*?)Resumo:', bloco, re.IGNORECASE | re.DOTALL)
                    if seccao_produtos_match:
                        texto_produtos = seccao_produtos_match.group(1)
                        linhas_produtos = padrao_produtos.findall(texto_produtos)
                        for linha in linhas_produtos:
                            produto = {
                                'cod': linha[0].strip(),
                                'produto': linha[1].replace('\n', ' ').strip(),
                                'os': linha[2].strip(),
                                'dimensoes': linha[3].strip(),
                                'qtde': linha[4].strip(),
                                'm2': linha[5].strip()
                            }
                            produtos.append(produto)
                    
                    if dados_cabecalho:
                        relatorios_extraidos.append({
                            'cabecalho': dados_cabecalho,
                            'produtos': produtos
                        })

                if not relatorios_extraidos:
                    flash('Nenhum dado de pedido válido foi encontrado. O formato do PDF pode ser diferente do esperado.', 'warning')
                    return render_template('relatorio_romaneio.html')

                flash(f'{len(relatorios_extraidos)} pedido(s) processado(s) com sucesso!', 'success')
                return render_template('relatorio_romaneio.html', relatorios=relatorios_extraidos)

            except Exception as e:
                flash(f'Ocorreu um erro inesperado ao processar o PDF: {e}', 'danger')
                return redirect(request.url)

    return render_template('relatorio_romaneio.html')

@app.route('/relatorio-pedidos')
@login_required
def relatorio_pedidos():
    """Exibe a página da nova ferramenta de Relatório de Pedidos."""
    # Por enquanto, esta função apenas renderiza a página.
    # A lógica de upload e análise de PDF será adicionada depois.
    return render_template('relatorio_pedidos.html')

@app.route('/cadastro-clientes')
@login_required
def cadastro_clientes():
    """Exibe a página de gestão de clientes com paginação e filtros (30 clientes por página)."""
    page = request.args.get('page', 1, type=int)
    per_page = 30  # Limite de 30 clientes por página
    
    # Parâmetros de filtro
    search_name = request.args.get('search_name', '').strip()
    filter_tipo_pessoa = request.args.get('filter_tipo_pessoa', '')
    search_cpf_cnpj = request.args.get('search_cpf_cnpj', '').strip()
    
    # Query base
    query = Cliente.query
    
    # Aplicar filtros
    if search_name:
        query = query.filter(Cliente.nome.ilike(f'%{search_name}%'))
    
    if filter_tipo_pessoa:
        query = query.filter(Cliente.tipo_pessoa == filter_tipo_pessoa)
    
    if search_cpf_cnpj:
        query = query.filter(Cliente.cpf_cnpj.ilike(f'%{search_cpf_cnpj}%'))
    
    # Paginação
    clientes_paginados = query.order_by(Cliente.numero_cliente).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('cadastro_clientes.html', 
                         clientes=clientes_paginados,
                         search_name=search_name,
                         filter_tipo_pessoa=filter_tipo_pessoa,
                         search_cpf_cnpj=search_cpf_cnpj)

@app.route('/novo-cliente/<tipo>', methods=['GET', 'POST'])
@login_required
def novo_cliente(tipo):
    """Renderiza o formulário e, ao salvar, valida nomes e CPFs duplicados."""
    if tipo not in ['rapido', 'completo']:
        return "Tipo de cadastro inválido", 404

    if request.method == 'POST':
        nome_form = request.form.get('nome')
        cpf_cnpj_form = request.form.get('cpf_cnpj') if request.form.get('cpf_cnpj') else None

        # --- NOVA VALIDAÇÃO DE NOME (CASE-INSENSITIVE) ---
        cliente_existente_nome = Cliente.query.filter(Cliente.nome.ilike(nome_form)).first()
        if cliente_existente_nome:
            flash(f'O cliente "{nome_form}" já está cadastrado.', 'danger')
            return render_template('novo_cliente.html', tipo=tipo, form_data=request.form)
        # --- FIM DA VALIDAÇÃO DE NOME ---
        
        if cpf_cnpj_form:
            cliente_existente_cpf = Cliente.query.filter_by(cpf_cnpj=cpf_cnpj_form).first()
            if cliente_existente_cpf:
                flash('Este CPF/CNPJ já está cadastrado.', 'danger')
                return render_template('novo_cliente.html', tipo=tipo, form_data=request.form)

        ultimo_cliente = Cliente.query.order_by(Cliente.numero_cliente.desc()).first()
        novo_numero = (ultimo_cliente.numero_cliente + 1) if ultimo_cliente else 1
        
        novo_cliente_obj = Cliente(
            numero_cliente=novo_numero,
            nome=nome_form,
            telefone=request.form.get('telefone') if request.form.get('telefone') else None,
            tipo_pessoa=request.form.get('tipo_pessoa'),
            cpf_cnpj=cpf_cnpj_form,
            como_conheceu=request.form.get('como_conheceu') if request.form.get('como_conheceu') else None,
            rua=request.form.get('rua') if request.form.get('rua') else None,
            numero_endereco=request.form.get('numero_endereco') if request.form.get('numero_endereco') else None,
            complemento=request.form.get('complemento') if request.form.get('complemento') else None,
            bairro=request.form.get('bairro') if request.form.get('bairro') else None,
            cidade=request.form.get('cidade') if request.form.get('cidade') else None,
            uf=request.form.get('uf') if request.form.get('uf') else None,
            cep=request.form.get('cep') if request.form.get('cep') else None,
            observacoes=request.form.get('observacoes') if request.form.get('observacoes') else None
        )
        db.session.add(novo_cliente_obj)
        db.session.commit()
        
        # Lógica para atualizar entradas com nome temporário
        entradas_para_atualizar = Entrada.query.filter_by(cliente_nome_temp=novo_cliente_obj.nome).all()
        if entradas_para_atualizar:
            for entrada in entradas_para_atualizar:
                entrada.cliente_id = novo_cliente_obj.id
                entrada.cliente_nome_temp = None
            db.session.commit()
            flash(f'Cliente "{novo_cliente_obj.nome}" cadastrado e {len(entradas_para_atualizar)} entrada(s) atualizada(s) com sucesso!', 'success')
        else:
            flash(f'Cliente "{novo_cliente_obj.nome}" cadastrado com sucesso!', 'success')
        
        return redirect(url_for('cadastro_clientes'))

    nome_preenchido = request.args.get('nome', '')
    form_data = {'nome': nome_preenchido}

    return render_template('novo_cliente.html', tipo=tipo, form_data=form_data)

@app.route('/editar-cliente/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_cliente(id):
    """Renderiza o formulário para editar, com validação de nome e CPF/CNPJ duplicados."""
    cliente = Cliente.query.get_or_404(id)
    if request.method == 'POST':
        nome_form = request.form.get('nome')
        cpf_cnpj_form = request.form.get('cpf_cnpj') if request.form.get('cpf_cnpj') else None

        # --- NOVA VALIDAÇÃO DE NOME (CASE-INSENSITIVE) ---
        # Verifica se já existe OUTRO cliente com o mesmo nome
        cliente_existente_nome = Cliente.query.filter(Cliente.id != id, Cliente.nome.ilike(nome_form)).first()
        if cliente_existente_nome:
            flash(f'Já existe outro cliente cadastrado com o nome "{nome_form}".', 'danger')
            return render_template('editar_cliente.html', cliente=cliente)
        # --- FIM DA VALIDAÇÃO DE NOME ---

        if cpf_cnpj_form and cpf_cnpj_form != cliente.cpf_cnpj:
            cliente_existente_cpf = Cliente.query.filter_by(cpf_cnpj=cpf_cnpj_form).first()
            if cliente_existente_cpf:
                flash('Este CPF/CNPJ já pertence a outro cliente.', 'danger')
                return render_template('editar_cliente.html', cliente=cliente)

        # Atualiza o objeto cliente
        cliente.nome = nome_form
        cliente.telefone = request.form.get('telefone') if request.form.get('telefone') else None
        cliente.tipo_pessoa = request.form.get('tipo_pessoa')
        cliente.cpf_cnpj = cpf_cnpj_form
        cliente.como_conheceu = request.form.get('como_conheceu') if request.form.get('como_conheceu') else None
        cliente.rua = request.form.get('rua') if request.form.get('rua') else None
        cliente.numero_endereco = request.form.get('numero_endereco') if request.form.get('numero_endereco') else None
        cliente.complemento = request.form.get('complemento') if request.form.get('complemento') else None
        cliente.bairro = request.form.get('bairro') if request.form.get('bairro') else None
        cliente.cidade = request.form.get('cidade') if request.form.get('cidade') else None
        cliente.uf = request.form.get('uf') if request.form.get('uf') else None
        cliente.cep = request.form.get('cep') if request.form.get('cep') else None
        cliente.observacoes = request.form.get('observacoes') if request.form.get('observacoes') else None
        
        db.session.commit()
        flash(f'Dados do cliente "{cliente.nome}" atualizados com sucesso!', 'success')
        return redirect(url_for('cadastro_clientes'))

    return render_template('editar_cliente.html', cliente=cliente)

@app.route('/excluir-cliente/<int:id>', methods=['POST'])
@login_required
def excluir_cliente(id):
    """Exclui um cliente do banco de dados."""
    cliente = Cliente.query.get_or_404(id)
    try:
        db.session.delete(cliente)
        db.session.commit()
        flash(f'Cliente "{cliente.nome}" excluído com sucesso.', 'success')
    except IntegrityError:
        db.session.rollback()
        flash(f'Não foi possível excluir o cliente "{cliente.nome}" pois ele está associado a um ou mais pedidos/orçamentos.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Ocorreu um erro ao excluir o cliente: {e}', 'danger')
        
    return redirect(url_for('cadastro_clientes'))

@app.route('/exportar-clientes')
@login_required
def exportar_clientes():
    """Exporta todos os clientes cadastrados para um ficheiro .xlsx."""
    try:
        import openpyxl
        from flask import Response
        import io

        clientes = Cliente.query.order_by(Cliente.numero_cliente).all()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Clientes"

        # Define os cabeçalhos das colunas
        headers = [
            'N° Cliente', 'Nome', 'Telefone', 'Tipo Pessoa', 'CPF/CNPJ', 
            'Como Conheceu', 'Rua', 'Número', 'Complemento', 'Bairro', 
            'Cidade', 'UF', 'CEP', 'Observações'
        ]
        sheet.append(headers)

        # Adiciona os dados de cada cliente
        for cliente in clientes:
            sheet.append([
                cliente.numero_cliente,
                cliente.nome,
                cliente.telefone,
                cliente.tipo_pessoa,
                cliente.cpf_cnpj,
                cliente.como_conheceu,
                cliente.rua,
                cliente.numero_endereco,
                cliente.complemento,
                cliente.bairro,
                cliente.cidade,
                cliente.uf,
                cliente.cep,
                cliente.observacoes
            ])

        # Cria um arquivo em memória para não salvar no disco do servidor
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        # Retorna o arquivo para download no navegador
        return Response(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": "attachment;filename=cadastro_clientes.xlsx"})
    except Exception as e:
        flash(f"Ocorreu um erro ao exportar os dados: {e}", "danger")
        return redirect(url_for('cadastro_clientes'))


@app.route('/importar-clientes', methods=['POST'])
@login_required
def importar_clientes():
    """Importa novos clientes a partir de um ficheiro .xlsx."""
    if 'xlsx_file' not in request.files:
        flash('Nenhum ficheiro selecionado.', 'warning')
        return redirect(url_for('cadastro_clientes'))

    ficheiro = request.files['xlsx_file']

    if ficheiro.filename == '':
        flash('Nenhum ficheiro selecionado.', 'warning')
        return redirect(url_for('cadastro_clientes'))

    if ficheiro and ficheiro.filename.lower().endswith('.xlsx'):
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(ficheiro)
            sheet = workbook.active

            clientes_adicionados = 0
            clientes_ignorados = 0
            
            ultimo_cliente = Cliente.query.order_by(Cliente.numero_cliente.desc()).first()
            proximo_numero = (ultimo_cliente.numero_cliente + 1) if ultimo_cliente else 1

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or all(cell is None for cell in row) or not row[0]:
                    continue # Pula linhas vazias ou sem nome

                # Mapeia as colunas para os campos do modelo
                nome, telefone, tipo_pessoa, cpf_cnpj, como_conheceu, rua, numero_endereco, complemento, bairro, cidade, uf, cep, observacoes = (row + (None,) * 13)[:13]

                # Validação: Ignora se já existir um cliente com o mesmo CPF/CNPJ (se houver um)
                if cpf_cnpj:
                    cliente_existente = Cliente.query.filter_by(cpf_cnpj=cpf_cnpj).first()
                    if cliente_existente:
                        clientes_ignorados += 1
                        continue
                
                novo_cliente = Cliente(
                    numero_cliente=proximo_numero,
                    nome=nome,
                    telefone=telefone,
                    tipo_pessoa=tipo_pessoa if tipo_pessoa in ['Física', 'Jurídica'] else 'Física',
                    cpf_cnpj=cpf_cnpj,
                    como_conheceu=como_conheceu,
                    rua=rua,
                    numero_endereco=numero_endereco,
                    complemento=complemento,
                    bairro=bairro,
                    cidade=cidade,
                    uf=uf,
                    cep=cep,
                    observacoes=observacoes
                )
                db.session.add(novo_cliente)
                clientes_adicionados += 1
                proximo_numero += 1

            if clientes_adicionados > 0:
                db.session.commit()

            mensagem = f'{clientes_adicionados} cliente(s) importado(s) com sucesso!'
            if clientes_ignorados > 0:
                mensagem += f' {clientes_ignorados} foram ignorados por já possuírem CPF/CNPJ cadastrado.'

            flash(mensagem, 'success' if clientes_adicionados > 0 else 'info')

        except Exception as e:
            db.session.rollback()
            flash(f'Ocorreu um erro ao processar o ficheiro: {e}', 'danger')

        return redirect(url_for('cadastro_clientes'))

    flash('Formato de ficheiro inválido. Por favor, envie um ficheiro .xlsx.', 'danger')
    return redirect(url_for('cadastro_clientes'))

@app.route('/exportar-painel')
@login_required
def exportar_painel():
    import openpyxl
    from flask import Response
    import io

    # Seleciona apenas as entradas ativas (não arquivadas)
    entradas = Entrada.query.filter_by(arquivado=False).order_by(Entrada.numero_pedido).all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Painel_Controle"

    # Cabeçalhos
    sheet.append([
        'N° Pedido', 'Tipo', 'Cliente', 'Obra', 'Status', 
        'Descrição', 'Observações', 'Data de Registro'
    ])

    # Dados
    for entrada in entradas:
        # Define o nome do cliente (seja o cadastrado ou o temporário)
        nome_cliente = entrada.cliente.nome if entrada.cliente else entrada.cliente_nome_temp
        
        sheet.append([
            entrada.numero_pedido,
            entrada.tipo,
            nome_cliente,
            entrada.obra,
            entrada.status,
            entrada.descricao,
            entrada.observacoes,
            entrada.data_registro.strftime('%Y-%m-%d %H:%M:%S')
        ])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment;filename=painel_controle.xlsx"})

@app.route('/exportar-arquivados')
@login_required
def exportar_arquivados():
    import openpyxl
    from flask import Response
    import io

    # Seleciona apenas as entradas arquivadas
    entradas = Entrada.query.filter_by(arquivado=True).order_by(Entrada.numero_pedido).all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Arquivados"

    # Cabeçalhos
    sheet.append([
        'N° Pedido', 'Tipo', 'Cliente', 'Obra', 'Status', 
        'Descrição', 'Observações', 'Data de Registro'
    ])

    # Dados
    for entrada in entradas:
        nome_cliente = entrada.cliente.nome if entrada.cliente else entrada.cliente_nome_temp
        
        sheet.append([
            entrada.numero_pedido,
            entrada.tipo,
            nome_cliente,
            entrada.obra,
            entrada.status,
            entrada.descricao,
            entrada.observacoes,
            entrada.data_registro.strftime('%Y-%m-%d %H:%M:%S')
        ])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return Response(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment;filename=arquivados.xlsx"})

@app.route('/importar-entradas', methods=['POST'])
@login_required
def importar_entradas():
    if 'xlsx_file' not in request.files:
        flash('Nenhum ficheiro selecionado.', 'warning')
        return redirect(request.referrer)

    ficheiro = request.files['xlsx_file']

    if ficheiro.filename == '':
        flash('Nenhum ficheiro selecionado.', 'warning')
        return redirect(request.referrer)

    if ficheiro and ficheiro.filename.lower().endswith('.xlsx'):
        try:
            workbook = openpyxl.load_workbook(ficheiro)
            sheet = workbook.active

            entradas_adicionadas = 0
            entradas_ignoradas = 0
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # --- INÍCIO DA CORREÇÃO ---
                # Pula a iteração se a linha for nula ou inteiramente vazia
                if not row or all(cell is None for cell in row):
                    continue
                # --- FIM DA CORREÇÃO ---
                
                # Garante que a linha tenha colunas suficientes antes de tentar acessá-las
                if len(row) < 7:
                    continue
                
                num_pedido_str, tipo_entrada, nome_cliente, obra, status, descricao, observacoes = row[:7]

                if not num_pedido_str or not tipo_entrada or not nome_cliente or not descricao:
                    continue

                try:
                    num_pedido = int(num_pedido_str)
                except (ValueError, TypeError):
                    continue

                if Entrada.query.filter_by(numero_pedido=num_pedido).first():
                    entradas_ignoradas += 1
                    continue

                cliente_obj = Cliente.query.filter(Cliente.nome.ilike(str(nome_cliente))).first()

                nova_entrada = Entrada(
                    numero_pedido=num_pedido,
                    tipo=tipo_entrada,
                    cliente_id=cliente_obj.id if cliente_obj else None,
                    cliente_nome_temp=str(nome_cliente) if not cliente_obj else None,
                    obra=str(obra) if obra else None,
                    status=str(status) if status in ['Não iniciado', 'Em andamento', 'Concluído'] else 'Não iniciado',
                    descricao=str(descricao),
                    observacoes=str(observacoes) if observacoes else None
                )
                db.session.add(nova_entrada)
                entradas_adicionadas += 1

            if entradas_adicionadas > 0:
                db.session.commit()
                global last_update_time
                last_update_time = datetime.now()

            mensagem = ""
            if entradas_adicionadas > 0:
                mensagem += f'{entradas_adicionadas} entrada(s) importada(s) com sucesso! '
            if entradas_ignoradas > 0:
                mensagem += f'{entradas_ignoradas} entrada(s) foram ignoradas por já existirem.'

            if mensagem:
                flash(mensagem, 'success' if entradas_adicionadas > 0 else 'info')
            else:
                flash('Nenhum dado novo para importar foi encontrado no ficheiro.', 'info')

        except Exception as e:
            db.session.rollback()
            flash(f'Ocorreu um erro ao processar o ficheiro: {e}', 'danger')

        return redirect(request.referrer)

    flash('Formato de ficheiro inválido. Por favor, envie um ficheiro .xlsx.', 'danger')
    return redirect(request.referrer)

@app.route('/buscar-clientes')
@login_required
def buscar_clientes():
    """Busca clientes por nome para o autocomplete."""
    search = request.args.get('term', '')
    query = Cliente.query.filter(Cliente.nome.ilike(f'%{search}%')).limit(10).all()
    # --- ALTERAÇÃO AQUI ---
    # Removemos o "N°: " da formatação da 'label'
    results = [{'id': cliente.id, 'label': f"{cliente.nome} ({cliente.numero_cliente})", 'value': cliente.nome} for cliente in query]
    return jsonify(results)

# @app.cli.command("init-db")
# def init_db_command():
#    db.create_all()
#    print("Base de dados inicializada e tabelas criadas.")


# @app.cli.command("create-admin")
# def create_admin_command():
#    username = input("Digite o nome de utilizador do ADMIN: ")
#    password = getpass("Digite a senha do ADMIN: ")
#    existing_user = User.query.filter_by(username=username).first()
 #   if existing_user:
 #       print(f"Erro: O utilizador '{username}' já existe.")
 #       return
 #   hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
 #   new_user = User(username=username, password_hash=hashed_password, is_admin=True)
 #   db.session.add(new_user)
 #   db.session.commit()
 #   print(f"Utilizador ADMINISTRADOR '{username}' criado com sucesso!")

if __name__ == '__main__':
    # Usa o iniciador do SocketIO, que é compatível com eventlet e ativa o modo de depuração
    # socketio.run(app, debug=True, host='127.0.0.1', port=5000)
    app.run(debug=True, host='localhost', port=5000)